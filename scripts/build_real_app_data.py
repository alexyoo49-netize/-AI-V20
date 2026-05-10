import csv
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


BASE = Path.cwd()
USER_ROOT = Path.home() / "Desktop" / "데이터 셋 모음 고용노동"
OUT_JSON = BASE / "public" / "data" / "datasets.json"
SUMMARY_JSON = BASE / "public" / "data" / "real_data_summary.json"
WAGEWORK_RAW = BASE / "outputs" / "wagework" / "wagework_selected_wage_results_raw.json"


REGION_ALIASES = {
    "서울": ["서울", "서울특별시"],
    "부산": ["부산", "부산광역시"],
    "대구": ["대구", "대구광역시"],
    "인천": ["인천", "인천광역시"],
    "광주": ["광주", "광주광역시"],
    "대전": ["대전", "대전광역시"],
    "울산": ["울산", "울산광역시"],
    "세종": ["세종", "세종특별자치시"],
    "경기": ["경기", "경기도"],
    "강원": ["강원", "강원특별자치도", "강원도"],
    "충북": ["충북", "충청북도"],
    "충남": ["충남", "충청남도"],
    "전북": ["전북", "전라북도", "전북특별자치도"],
    "전남": ["전남", "전라남도"],
    "경북": ["경북", "경상북도"],
    "경남": ["경남", "경상남도"],
    "제주": ["제주", "제주도", "제주특별자치도"],
}
REGIONS = list(REGION_ALIASES)


JOB_TERMS = {
    "데이터 분석": ["데이터", "통계", "빅데이터", "네트워크"],
    "사무": ["사무", "행정", "경영", "회계", "경리", "비서", "고객 상담"],
    "마케팅": ["마케팅", "광고", "홍보", "상품 기획", "조사 전문가"],
    "개발": ["소프트웨어", "프로그래머", "웹", "응용", "시스템"],
    "디자인": ["디자인", "디자이너", "시각"],
}

NCS_DETAIL_TERMS = {
    "데이터 분석": ["빅데이터분석", "인공지능"],
    "사무": ["총무", "인사", "회계", "경영기획"],
    "마케팅": ["마케팅전략기획", "고객관리", "통계조사"],
    "개발": ["응용SW엔지니어링", "DB엔지니어링", "UI/UX엔지니어링", "인공지능플랫폼구축"],
    "디자인": ["디지털디자인", "시각디자인", "제품디자인", "서비스경험디자인", "VR콘텐츠디자인"],
}

NCS_LARGE_TERMS = {
    "데이터 분석": ["정보통신"],
    "개발": ["정보통신"],
    "사무": ["경영"],
    "마케팅": ["경영"],
    "디자인": ["디자인"],
}


JOB_META = {
    "데이터 분석": {
        "ncs": "정보기술 > 빅데이터분석",
        "required": ["Excel", "SQL", "Python", "데이터 시각화", "기초 통계", "문제정의"],
        "adjacent": ["데이터 운영", "BI 리포팅", "사무자동화", "마케팅 데이터 보조"],
        "wageCodes": ["223"],
    },
    "사무": {
        "ncs": "경영, 회계, 사무 > 총무, 인사, 회계",
        "required": ["Excel", "문서작성", "커뮤니케이션", "일정관리", "자료정리"],
        "adjacent": ["영업지원", "HR 운영", "총무", "고객운영"],
        "wageCodes": ["312", "313", "314", "399"],
    },
    "마케팅": {
        "ncs": "경영, 회계, 사무 > 마케팅",
        "required": ["콘텐츠 기획", "Excel", "데이터 리터러시", "SNS 운영", "문서작성"],
        "adjacent": ["콘텐츠 운영", "CRM 보조", "광고 운영", "영업지원"],
        "wageCodes": ["273"],
    },
    "개발": {
        "ncs": "정보기술 > 응용SW엔지니어링",
        "required": ["JavaScript", "Python", "Git", "SQL", "API 이해", "문제해결"],
        "adjacent": ["QA", "웹 퍼블리싱", "데이터 엔지니어링 보조", "IT 운영"],
        "wageCodes": ["222"],
    },
    "디자인": {
        "ncs": "문화, 예술, 디자인 > 디자인",
        "required": ["Figma", "포트폴리오", "사용자 이해", "시각디자인", "커뮤니케이션"],
        "adjacent": ["콘텐츠 디자인", "서비스 운영", "UX 리서치 보조", "마케팅 디자인"],
        "wageCodes": ["285"],
    },
}


def short_region(text):
    text = str(text or "")
    for region, aliases in REGION_ALIASES.items():
        if any(alias in text for alias in aliases):
            return region
    return None


def parse_num(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", ""))
    except ValueError:
        return 0.0


def parse_month(value):
    text = str(value or "")
    m = re.search(r"(20\d{2}).*?(\d{1,2})", text)
    if not m:
        return None
    return f"{m.group(1)}-{int(m.group(2)):02d}"


def normalize(value, min_v, max_v, lo=0.0, hi=1.0):
    if max_v <= min_v:
        return (lo + hi) / 2
    return lo + (value - min_v) / (max_v - min_v) * (hi - lo)


def find_file(name_contains, parent_contains=None, suffix=None):
    for p in USER_ROOT.rglob("*"):
        if not p.is_file():
            continue
        if suffix and p.suffix.lower() != suffix:
            continue
        if name_contains not in p.name:
            continue
        if parent_contains and parent_contains not in p.parent.name:
            continue
        return p
    raise FileNotFoundError(name_contains)


def find_files(name_contains, parent_contains=None, suffix=None):
    rows = []
    for p in USER_ROOT.rglob("*"):
        if not p.is_file():
            continue
        if suffix and p.suffix.lower() != suffix:
            continue
        if name_contains not in p.name:
            continue
        if parent_contains and parent_contains not in p.parent.name:
            continue
        rows.append(p)
    return sorted(rows, key=lambda p: p.name)


def iter_region_occupation_rows():
    files = find_files("구인구직취업현황", parent_contains="지역별", suffix=".xlsx")
    latest_by_region = {}
    for p in files:
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        region = short_region(ws.cell(9, 2).value) or short_region(p.name)
        wb.close()
        if not region:
            continue
        previous = latest_by_region.get(region)
        if previous is None or p.name > previous.name:
            latest_by_region[region] = p

    for p in sorted(latest_by_region.values(), key=lambda item: item.name):
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        region = short_region(ws.cell(9, 2).value) or short_region(p.name)
        current_month = None
        current_mid = None
        current_small = None
        for row in ws.iter_rows(min_row=15, values_only=True):
            if row[0] not in (None, ""):
                current_month = parse_month(row[0])
            if row[1] not in (None, ""):
                current_mid = str(row[1])
            if row[2] not in (None, ""):
                current_small = str(row[2])
            if not region or not current_month:
                continue
            openings, seekers, employed = parse_num(row[3]), parse_num(row[4]), parse_num(row[5])
            if openings + seekers + employed <= 0:
                continue
            label = " ".join([current_mid or "", current_small or ""]).strip()
            if not label or "총계" in label:
                continue
            yield {
                "region": region,
                "month": current_month,
                "occupation": label,
                "openings": openings,
                "seekers": seekers,
                "employed": employed,
            }


def build_labor_market():
    total = defaultdict(lambda: defaultdict(lambda: {"openings": 0.0, "seekers": 0.0, "employed": 0.0}))
    by_job = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {"openings": 0.0, "seekers": 0.0, "employed": 0.0})))
    for row in iter_region_occupation_rows():
        region, month = row["region"], row["month"]
        for key in ("openings", "seekers", "employed"):
            total[region][month][key] += row[key]
        for job, terms in JOB_TERMS.items():
            if any(term in row["occupation"] for term in terms):
                for key in ("openings", "seekers", "employed"):
                    by_job[region][job][month][key] += row[key]

    all_months = sorted({m for region in total.values() for m in region})
    recent_months = all_months[-12:]
    region_stats = {}
    job_stats = defaultdict(dict)
    for region in REGIONS:
        agg = {"openings": 0.0, "seekers": 0.0, "employed": 0.0}
        for month in recent_months:
            for key in agg:
                agg[key] += total[region][month][key]
        agg["openingsPerSeeker"] = agg["openings"] / agg["seekers"] if agg["seekers"] else 0
        agg["employmentPerSeeker"] = agg["employed"] / agg["seekers"] if agg["seekers"] else 0
        region_stats[region] = agg

        for job in JOB_META:
            j = {"openings": 0.0, "seekers": 0.0, "employed": 0.0}
            for month in recent_months:
                for key in j:
                    j[key] += by_job[region][job][month][key]
            j["openingsPerSeeker"] = j["openings"] / j["seekers"] if j["seekers"] else 0
            j["employmentPerSeeker"] = j["employed"] / j["seekers"] if j["seekers"] else 0
            job_stats[region][job] = j
    return recent_months, region_stats, job_stats


def read_wide_region_file(file_part, data_row_start, month_row, metric_name=None):
    p = find_file(file_part, suffix=".xlsx")
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    months = [parse_month(ws.cell(month_row, col).value) for col in range(2, ws.max_column + 1)]
    data = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=data_row_start, values_only=True):
        region = short_region(row[0])
        if not region:
            continue
        for idx, value in enumerate(row[1:], 0):
            month = months[idx] if idx < len(months) else None
            if month:
                data[region][month] += parse_num(value)
    return p.name, data


def build_training_access(recent_months):
    sources = {}
    try:
        name, cards = read_wide_region_file("내일배움카드발급현황", 15, 12)
        sources["내일배움카드 발급"] = name
    except Exception:
        cards = defaultdict(lambda: defaultdict(float))
    try:
        name, training = read_wide_region_file("실업자훈련실시현황", 15, 12)
        sources["실업자훈련 실시"] = name
    except Exception:
        training = defaultdict(lambda: defaultdict(float))

    metrics = {}
    totals = []
    for region in REGIONS:
        card_sum = sum(cards[region].get(m, 0) for m in recent_months)
        train_sum = sum(training[region].get(m, 0) for m in recent_months)
        metrics[region] = {"cardIssued": card_sum, "trainingParticipants": train_sum}
        totals.append(card_sum + train_sum)
    min_v, max_v = min(totals or [0]), max(totals or [1])
    for region in REGIONS:
        v = metrics[region]["cardIssued"] + metrics[region]["trainingParticipants"]
        metrics[region]["trainingAccess"] = round(normalize(v, min_v, max_v, 0.35, 0.9), 3)
    return metrics, sources


def count_agencies():
    counts = {region: Counter() for region in REGIONS}
    sources = {}

    try:
        p = find_file("청년일자리도약장려금", suffix=".csv")
        sources["청년일자리도약장려금 운영기관"] = p.name
        with p.open(encoding="cp949", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                region = short_region(" ".join(str(v) for v in row.values()))
                if region:
                    counts[region]["subsidyOperators"] += 1
    except Exception:
        pass

    try:
        p = find_file("국민취업지원제도 운영기관", suffix=".xlsx")
        sources["국민취업지원제도 운영기관"] = p.name
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=12, values_only=True):
            region = short_region(" ".join(str(v or "") for v in row))
            if region:
                counts[region]["counselingOperators"] += 1
    except Exception:
        pass

    return counts, sources


def load_wagework():
    by_code = {}
    if not WAGEWORK_RAW.exists():
        return by_code
    raw = json.loads(WAGEWORK_RAW.read_text(encoding="utf-8"))
    for rec in raw:
        result = ((rec.get("result") or {}).get("data") or {}).get("resultList")
        if not isinstance(result, dict):
            continue
        code = str(rec.get("occpClCd"))
        by_code[code] = {
            "name": rec.get("occpCfnm"),
            "avgAnnualKrwThousand": parse_num(result.get("entrAvrgAnslAmt")),
            "p25AnnualKrwThousand": parse_num(result.get("lwprAvrgIcomAmt")),
            "medianAnnualKrwThousand": parse_num(result.get("midAvrgIcomAmt")),
            "p75AnnualKrwThousand": parse_num(result.get("upprAvrgIcomAmt")),
        }
    return by_code


def build_ncs_units():
    units_by_job = defaultdict(list)
    try:
        p = find_file("NCS", suffix=".xlsx")
    except FileNotFoundError:
        return {}, None
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        large, mid, small, detail, code, name, level = row[1], row[3], row[5], row[7], row[8], row[9], row[10]
        target_text = " ".join(str(v or "") for v in (detail, name))
        large_text = str(large or "")
        for job, terms in NCS_DETAIL_TERMS.items():
            if not any(term in large_text for term in NCS_LARGE_TERMS.get(job, [])):
                continue
            if any(term in target_text for term in terms):
                units_by_job[job].append(
                    {
                        "code": code,
                        "name": name,
                        "level": level,
                        "category": " > ".join(str(v) for v in (large, mid, small, detail) if v),
                    }
                )
                break
    for job in units_by_job:
        units_by_job[job] = units_by_job[job][:12]
    return dict(units_by_job), p.name


def build_job_profiles(job_region_stats):
    wage_by_code = load_wagework()
    ncs_units, ncs_source = build_ncs_units()
    profiles = {}
    for job, meta in JOB_META.items():
        selected = [wage_by_code[c] for c in meta["wageCodes"] if c in wage_by_code]
        if selected:
            p25_months = [w["p25AnnualKrwThousand"] / 12 / 10 for w in selected if w["p25AnnualKrwThousand"]]
            med_months = [w["medianAnnualKrwThousand"] / 12 / 10 for w in selected if w["medianAnnualKrwThousand"]]
            avg_months = [w["avgAnnualKrwThousand"] / 12 / 10 for w in selected if w["avgAnnualKrwThousand"]]
            p25_month = sum(p25_months) / len(p25_months) if p25_months else 260
            median_month = sum(med_months) / len(med_months) if med_months else p25_month
            avg_month = sum(avg_months) / len(avg_months) if avg_months else median_month
            youth_adjusted = max(210, p25_month * 0.86)
        else:
            p25_month, median_month, avg_month, youth_adjusted = 240, 270, 300, 240

        # 직무별 훈련효과는 전국 직무 취업률과 NCS 매칭 유무를 이용해 보수적으로 보정한다.
        national = {"openings": 0.0, "seekers": 0.0, "employed": 0.0}
        for region in REGIONS:
            for key in national:
                national[key] += job_region_stats[region][job][key]
        placement = national["employed"] / national["seekers"] if national["seekers"] else 0.3
        training_boost = round(max(0.06, min(0.16, 0.06 + placement * 0.22 + (0.015 if ncs_units.get(job) else 0))), 3)

        profiles[job] = {
            "ncs": meta["ncs"],
            "required": meta["required"],
            "adjacent": meta["adjacent"],
            "baseWage": round(youth_adjusted),
            "trainingBoost": training_boost,
            "sourceIndustry": "워크피디아 직업분류 기준",
            "wageBasis": {
                "source": "임금직업포털 워크피디아 맞춤형 임금정보",
                "codes": meta["wageCodes"],
                "p25MonthlyManwon": round(p25_month, 1),
                "medianMonthlyManwon": round(median_month, 1),
                "averageMonthlyManwon": round(avg_month, 1),
                "youthAdjustedMonthlyManwon": round(youth_adjusted, 1),
                "note": "워크피디아는 2024년 기준 추정치이므로 청년 초입 경로에는 하위 25퍼센트 임금을 보수적으로 조정해 반영",
            },
            "ncsUnits": ncs_units.get(job, []),
        }
    return profiles, ncs_source


def build_dataset_catalog(source_files):
    return [
        {"id": "LOCAL-REGJOB", "title": "지역별 직종별 구인구직취업현황 월별 자료", "role": "지역과 직무별 구인, 구직, 취업 실적을 이용해 취업확률 기준선을 계산", "sourceFile": "지역별 직종별 구인구직취업현황(월) 폴더"},
        {"id": "LOCAL-WAGEWORK", "title": "임금직업포털 워크피디아 맞춤형 임금정보", "role": "직무별 기대임금 기준선과 보수적 청년 초입 임금값 산정", "sourceFile": "outputs/wagework"},
        {"id": "LOCAL-NCS", "title": "NCS 능력단위 목록 2025.12.16 고시 기준", "role": "희망직무와 보유역량 간 스킬갭 계산의 기준", "sourceFile": source_files.get("ncs")},
        {"id": "LOCAL-TRAIN", "title": "내일배움카드 및 실업자훈련 실시 현황", "role": "지역별 훈련 접근성과 훈련 경로 효과 보정", "sourceFile": ", ".join(source_files.get("training", []))},
        {"id": "LOCAL-COUNSEL", "title": "국민취업지원제도 운영기관 목록", "role": "상담기관 접근성과 상담 경로 효과 산정", "sourceFile": source_files.get("counseling")},
        {"id": "LOCAL-SUBSIDY", "title": "청년일자리도약장려금 사업 운영기관", "role": "장려금 활용 기업 경로의 지역 접근성 보정", "sourceFile": source_files.get("subsidy")},
    ]


def build_job_market(job_region_stats):
    job_market = {region: {} for region in REGIONS}
    for job in JOB_META:
        open_ratios = [job_region_stats[r][job]["openingsPerSeeker"] for r in REGIONS]
        emp_ratios = [job_region_stats[r][job]["employmentPerSeeker"] for r in REGIONS]
        comp_ratios = [
            job_region_stats[r][job]["seekers"] / max(job_region_stats[r][job]["openings"], 1)
            for r in REGIONS
        ]
        min_open, max_open = min(open_ratios), max(open_ratios)
        min_emp, max_emp = min(emp_ratios), max(emp_ratios)
        min_comp, max_comp = min(comp_ratios), max(comp_ratios)
        for region in REGIONS:
            stat = job_region_stats[region][job]
            demand = (
                0.55 * normalize(stat["openingsPerSeeker"], min_open, max_open, 0.25, 0.9)
                + 0.45 * normalize(stat["employmentPerSeeker"], min_emp, max_emp, 0.25, 0.85)
            )
            competition = normalize(
                stat["seekers"] / max(stat["openings"], 1),
                min_comp,
                max_comp,
                0.25,
                0.9,
            )
            job_market[region][job] = {
                "demand": round(demand, 3),
                "competition": round(competition, 3),
                "rawCounts": {
                    "openings": round(stat["openings"]),
                    "jobSeekers": round(stat["seekers"]),
                    "employed": round(stat["employed"]),
                    "openingsPerSeeker": round(stat["openingsPerSeeker"], 3),
                    "employmentPerSeeker": round(stat["employmentPerSeeker"], 3),
                },
            }
    return job_market


def detect_duplicate_region_stats(region_stats):
    grouped = defaultdict(list)
    for region, stat in region_stats.items():
        grouped[(round(stat["openings"]), round(stat["seekers"]), round(stat["employed"]))].append(region)
    return [
        {"regions": regions, "reason": "same_openings_seekers_employed"}
        for regions in grouped.values()
        if len(regions) > 1
    ]


def main():
    recent_months, region_stats, job_region_stats = build_labor_market()
    training_metrics, training_sources = build_training_access(recent_months)
    agency_counts, agency_sources = count_agencies()
    job_profiles, ncs_source = build_job_profiles(job_region_stats)

    openings_ratio = [region_stats[r]["openingsPerSeeker"] for r in REGIONS]
    employment_ratio = [region_stats[r]["employmentPerSeeker"] for r in REGIONS]
    support_raw = []
    for region in REGIONS:
        support_raw.append(
            training_metrics[region]["trainingAccess"]
            + agency_counts[region]["counselingOperators"] * 0.018
            + agency_counts[region]["subsidyOperators"] * 0.014
        )
    min_open, max_open = min(openings_ratio), max(openings_ratio)
    min_emp, max_emp = min(employment_ratio), max(employment_ratio)
    min_sup, max_sup = min(support_raw), max(support_raw)

    market = {}
    for idx, region in enumerate(REGIONS):
        stat = region_stats[region]
        demand = 0.55 * normalize(stat["openingsPerSeeker"], min_open, max_open, 0.32, 0.82) + 0.45 * normalize(stat["employmentPerSeeker"], min_emp, max_emp, 0.3, 0.78)
        competition = normalize(stat["seekers"] / max(stat["openings"], 1), min(region_stats[r]["seekers"] / max(region_stats[r]["openings"], 1) for r in REGIONS), max(region_stats[r]["seekers"] / max(region_stats[r]["openings"], 1) for r in REGIONS), 0.34, 0.72)
        support_access = normalize(support_raw[idx], min_sup, max_sup, 0.35, 0.88)
        market[region] = {
            "demand": round(demand, 3),
            "competition": round(competition, 3),
            "wage": 1.0,
            "supportAccess": round(support_access, 3),
            "rawCounts": {
                "recentMonths": recent_months,
                "openings": round(stat["openings"]),
                "jobSeekers": round(stat["seekers"]),
                "employed": round(stat["employed"]),
                "openingsPerSeeker": round(stat["openingsPerSeeker"], 3),
                "employmentPerSeeker": round(stat["employmentPerSeeker"], 3),
                "trainingParticipants": round(training_metrics[region]["trainingParticipants"]),
                "cardIssued": round(training_metrics[region]["cardIssued"]),
                "counselingOperators": agency_counts[region]["counselingOperators"],
                "subsidyOperators": agency_counts[region]["subsidyOperators"],
            },
        }

    counseling_total = sum(agency_counts[r]["counselingOperators"] for r in REGIONS)
    subsidy_total = sum(agency_counts[r]["subsidyOperators"] for r in REGIONS)
    policy_effects = {
        "counseling": round(0.035 + min(0.045, counseling_total / 6000), 3),
        "trainingBase": round(sum(p["trainingBoost"] for p in job_profiles.values()) / len(job_profiles), 3),
        "subsidy": round(0.045 + min(0.035, subsidy_total / 2500), 3),
        "integratedSynergy": 0.045,
        "distancePenalty": 0.025,
        "jobSearchPressure": 0.025,
    }

    source_files = {
        "ncs": ncs_source,
        "training": list(training_sources.values()),
        "counseling": agency_sources.get("국민취업지원제도 운영기관"),
        "subsidy": agency_sources.get("청년일자리도약장려금 운영기관"),
    }
    datasets = build_dataset_catalog(source_files)
    job_market = build_job_market(job_region_stats)
    duplicate_warnings = detect_duplicate_region_stats(region_stats)
    data = {
        "source": {
            "catalogFile": "사용자 제공 고용노동 데이터셋 폴더 및 워크피디아 수집자료",
            "totalCatalogRows": 224,
            "localSourceFiles": len(list(USER_ROOT.rglob("*.*"))),
            "selectedRows": len(datasets),
            "generatedAt": "2026-05-08",
            "rawDataLoaded": True,
            "rawDatasetCount": len(datasets),
            "rawManifest": "public/data/real_data_summary.json",
            "analysisSummary": "public/data/real_data_summary.json",
            "analysisGeneratedAt": "2026-05-08",
        },
        "datasets": datasets,
        "market": market,
        "jobMarket": job_market,
        "jobProfiles": job_profiles,
        "dataQuality": {
            "duplicateRegionWarnings": duplicate_warnings,
        },
        "policyEffects": policy_effects,
        "simulationSpec": {
            "agentsPerRun": 1000,
            "runs": 500,
            "months": 12,
            "method": "규칙 기반 ABM",
            "dataBasis": "최근 12개월 지역별 직종별 구인구직취업 실적, 기관 접근성, 훈련 접근성, 워크피디아 임금 기준선",
        },
    }
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "recentMonths": recent_months,
        "regionStats": region_stats,
        "jobRegionStats": job_region_stats,
        "trainingMetrics": training_metrics,
        "agencyCounts": {r: dict(agency_counts[r]) for r in REGIONS},
        "sourceFiles": source_files,
        "dataQuality": {
            "duplicateRegionWarnings": duplicate_warnings,
        },
        "policyEffects": policy_effects,
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT_JSON}")
    print(f"wrote {SUMMARY_JSON}")


if __name__ == "__main__":
    main()
